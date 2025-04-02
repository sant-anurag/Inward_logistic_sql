# Importing tkinter module with alias name "tk"
import tkinter as tk
import mysql.connector
import ctypes
import subprocess
# Importing specific classes and functions from tkinter module
import pandas as pd
from tkinter import Canvas  # Importing Canvas specifically for background image handling

from tkinter import (
    Tk, Label, Entry, Button, Frame, Text, Scrollbar, Listbox,
    Menu, PhotoImage, filedialog, messagebox, Canvas, constants
)
import sys
from tkinter import ttk, messagebox
import xlsxwriter
import re
from openpyxl import load_workbook
from pathlib import Path
from tkcalendar import DateEntry
from datetime import datetime
from PIL import Image, ImageTk

from tkinter import *
# Importing specific class from tkinter.messagebox module
from tkinter.messagebox import showinfo

# Importing docx module to work with Word documents
import docx

# Importing openpyxl module to work with Excel files
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Importing DateEntry class from tkcalendar module
from tkcalendar import DateEntry

# Importing random module for generating random values
import random

# Importing partial function from functools module
from functools import partial

# Importing shutil module for working with files and directories
import shutil

# Importing os module for interacting with the operating system
import os

# Importing pyautogui module for automating mouse and keyboard actions
import pyautogui

# Importing Image and ImageTk classes from PIL module
from PIL import Image, ImageTk

from openpyxl.styles import Font, PatternFill,Alignment, Border, Side

# defining fonts for usage in project
NORM_FONT = ('times new roman', 13, 'normal')
NORM_FONT_MEDIUM_HIGH = ('times new roman', 15, 'normal')
NORM_FONT_MEDIUM_LOW = ('times new roman', 14, 'normal')
TIMES_NEW_ROMAN_BIG = ('times new roman', 16, 'normal')
NORM_VERDANA_FONT = ('verdana', 10, 'normal')
BOLD_VERDANA_FONT = ('verdana', 11, 'normal')
LARGE_VERDANA_FONT = ('verdana', 13, 'normal')
XXL_FONT = ('times new roman', 25, 'normal')
XL_FONT = ('times new roman', 20, 'normal')
L_FONT = ('times new roman', 15, 'normal')
MAX_ENTRIES_FOR_DISPLAY = 6

database_columns = ("Inward No", "Return Type", "Benefit Type", "Date", "Time", "Gate Entry No",
                   "Invoice No", "PO No", "BOE No", "Return Date", "Return Time","Supplier", "Material", "Qty", "Department",
                   "Project", "TPL Name","Vehicle", "Received", "Authorized", "Security", "Remark", "TPL Remarks")

database_fields = [
            "Inward No", "Return Type", "Benefit Type", "Date", "Time", "Gate Entry No", "Invoice No *", "PO No",
            "BOE No", "Return Date", "Return Time","Name of the Supplier", "Material Description", "Quantity *", "Department", "Project_Name", "TPL_Name","Vehicle No",
            "Received Name", "Authorized Sign", "Security Sign", "Remark","TPL Remarks"
        ]

database_fields_load = [
            "S.No","Inward No", "Return Type", "Benefit Type", "Date", "Time", "Gate Entry No", "Invoice No", "PO No",
            "BOE No", "Return Date", "Return Time","Name of the Supplier", "Material Description", "Qty", "Department", "Project_Name", "TPL_Name","Vehicle No",
            "Received Name", "Authorized Sign", "Security Sign", "Remark","TPL Remarks"
        ]
database_columns = [
    "Inward_No INT",
    "Return_Type VARCHAR(255)",
    "Benefit_Type VARCHAR(255)",
    "Date DATE",
    "Time TIME",
    "Gate_Entry_No VARCHAR(255)",
    "Invoice_No VARCHAR(255)",
    "PO_No VARCHAR(255)",
    "BOE_No VARCHAR(255)",
    "Return_Date DATE",
    "Return_Time TIME",
    "Supplier VARCHAR(255)",
    "Material VARCHAR(255)",
    "Qty INT",
    "Department VARCHAR(255)",
    "Project VARCHAR(255)",
    "TPL_Name VARCHAR(255)",
    "Vehicle VARCHAR(255)",
    "Received VARCHAR(255)",
    "Authorized VARCHAR(255)",
    "Security VARCHAR(255)",
    "Remark TEXT",
    "TPL_Remarks TEXT"
]

column_names = [
    "Inward_No",
    "Return_Type",
    "Benefit_Type",
    "Date",
    "Time",
    "Gate_Entry_No",
    "Invoice_No",
    "PO_No",
    "BOE_No",
    "Return_Date",
    "Return_Time",
    "Supplier",
    "Material",
    "Qty",
    "Department",
    "Project",
    "TPL_Name",
    "Vehicle",
    "Received",
    "Authorized",
    "Security",
    "Remark",
    "TPL_Remarks"
]

serverdb_config = {
    'user': 'forvia',
    'password': 'password@123',
    'host': '10.170.140.103',
    'port': 3306,
    'database': 'logistic'
}